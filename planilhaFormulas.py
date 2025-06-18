import customtkinter as ctk
from tkinter import StringVar
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.protection import SheetProtection
from tkinter import messagebox
import os

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.protection import SheetProtection
from tkinter import messagebox
from openpyxl.worksheet.datavalidation import DataValidation

def formatarPlanilha():
    if not arquivos_selecionados:
        messagebox.showwarning("Aviso", "Nenhum arquivo selecionado.")
        return

    for caminho in arquivos_selecionados:
        try:
            wb = load_workbook(caminho)
            ws = wb.active

            # 1. Desbloquear TODAS as células
            for row in ws.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)

            # 2. Bloquear colunas A até H
            colunas_bloqueadas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            for col in colunas_bloqueadas:
                for row in range(1, ws.max_row + 1):
                    ws[f"{col}{row}"].protection = Protection(locked=True)

            # 3. Validação de dados na coluna I ("Posição (APTO/NÃO APTO)")
            dv_apto = DataValidation(
                type="list",
                formula1='"APTO,NÃO APTO"',
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True
            )
            ws.add_data_validation(dv_apto)
            for row in range(2, ws.max_row + 1):
                dv_apto.add(ws[f"I{row}"])

            # 4. Validação com fórmula na coluna J ("Valor Apto")
            for row in range(2, ws.max_row + 1):
                formula = f'=IF(I{row}="APTO",TRUE,FALSE)'
                dv_formula = DataValidation(
                    type="custom",
                    formula1=formula,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="NÃO APTO",
                    error="Se a coluna I for preenchida com \"NÃO APTO\", a coluna J NÃO deve ser preenchida.\nPara esse alerta sumir, apague o conteúdo da célula J preenchida incorretamente."

                )
                ws.add_data_validation(dv_formula)
                dv_formula.add(ws[f"J{row}"])

            # 5. Proteger a planilha com senha deo123
            ws.protection = SheetProtection(sheet=True, password='senha1234')

            wb.save(caminho)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar {os.path.basename(caminho)}:\n{e}")
            continue

    messagebox.showinfo("Concluído", "Formatação aplicada com sucesso!")

def selecionar_arquivos():
    global arquivos_selecionados
    arquivos = ctk.filedialog.askopenfilenames(
        filetypes=[("Excel Files", "*.xlsx")],
        multiple=True
    )
    if arquivos:
        arquivos_selecionados = arquivos
        nomes = [os.path.basename(arq) for arq in arquivos]
        arquivos_var.set("\n\n".join(nomes))
    else:
        arquivos_var.set("Nenhum arquivo selecionado")
        arquivos_selecionados = []



app = ctk.CTk()
app.title("Planilha")
app.geometry("300x300")
app.resizable(False, False)

tituloLabel = ctk.CTkLabel(app, text="Arquivo(s)", font=("Arial", 20)).pack(pady=20)

arquivos_var = StringVar()
arquivos_var.set("Nenhum arquivo selecionado")

arquivos_selecionados = []

ctk.CTkButton(
    app,
    text="Selecionar arquivo(s)",
    width=200,
    command=selecionar_arquivos
).pack(pady=10)

ctk.CTkLabel(
    app,
    textvariable=arquivos_var,
    width=200,
    justify="left"
).pack(pady=10)

ctk.CTkButton(app, text="Carregar", width=200, command=formatarPlanilha).pack(pady=10)
app.mainloop()